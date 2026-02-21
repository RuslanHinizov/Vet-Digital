"""
Excel template parser for VetDigital procedure act templates.

Parses the 4 Kazakhstan government Excel form templates:
1. Без названия (4).xls  — Main procedure act form
2. Без названия (3).xls  — Expendable materials table
3. Без названия (5).xls  — Animal registry (individual records)
4. Без названия (6).xls  — Owner signature block

Supports both .xls (xlrd) and .xlsx (openpyxl).
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed — .xlsx parsing unavailable")

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    logger.warning("xlrd not installed — .xls parsing unavailable")


# ─── Cell helpers ─────────────────────────────────────────────────────────────

def _cell_str(ws, row: int, col: int) -> str:
    """Get cell value as stripped string (openpyxl, 1-indexed)."""
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ""


def _cell_int(ws, row: int, col: int) -> Optional[int]:
    """Get cell value as integer."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _cell_float(ws, row: int, col: int) -> Optional[float]:
    """Get cell value as float."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    try:
        return float(str(val))
    except (ValueError, TypeError):
        return None


def _cell_date(ws, row: int, col: int) -> Optional[datetime]:
    """Get cell value as datetime."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    # Try common KZ date formats
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            continue
    return None


# ─── Template 1: Main act form (Без названия (4).xls) ────────────────────────

def parse_main_act_template(file_path: str | Path) -> dict:
    """
    Parse the main procedure act Excel template.

    Expected layout (approximate — adjust row/col for actual template):
    Row 1: Ministry header
    Row 3: Act number (B3), Date (D3)
    Row 5: Disease (B5), Settlement (D5)
    Row 7: Specialist (B7), Species (D7)
    Row 9: Owner name (B9), Owner IIN (D9)
    Row 11: Procedure type (B11)
    Row 14: Male count (B14), Female (C14), Young (D14), Total (E14)
    Row 17: Vaccine name (B17)
    Row 18: Manufacturer (B18)
    Row 19: Series (B19)
    Row 20: State control no (B20)
    Row 21: Injection method (B21)
    Row 22: Dose adult (B22), Dose young (C22)

    Returns dict matching ProcedureAct model fields.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required for .xlsx parsing")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Template file not found: {file_path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    result = {}

    # Act metadata
    result["act_number"] = _cell_str(ws, 3, 2) or None
    result["act_date"] = _cell_date(ws, 3, 4)
    result["disease_name"] = _cell_str(ws, 5, 2) or None
    result["settlement"] = _cell_str(ws, 5, 4) or None
    result["specialist_name"] = _cell_str(ws, 7, 2) or None
    result["species_name"] = _cell_str(ws, 7, 4) or None
    result["owner_name"] = _cell_str(ws, 9, 2) or None
    result["owner_iin"] = _cell_str(ws, 9, 4) or None

    # Procedure type (map from Russian label)
    proc_type_raw = _cell_str(ws, 11, 2).lower()
    result["procedure_type"] = _map_procedure_type(proc_type_raw)

    # Animal counts
    result["male_count"] = _cell_int(ws, 14, 2)
    result["female_count"] = _cell_int(ws, 14, 3)
    result["young_count"] = _cell_int(ws, 14, 4)
    result["total_vaccinated"] = _cell_int(ws, 14, 5)

    # Drug information
    result["vaccine_name"] = _cell_str(ws, 17, 2) or None
    result["manufacturer"] = _cell_str(ws, 18, 2) or None
    result["series"] = _cell_str(ws, 19, 2) or None
    result["state_control_no"] = _cell_str(ws, 20, 2) or None
    result["injection_method"] = _cell_str(ws, 21, 2) or None
    result["dose_adult_ml"] = _cell_float(ws, 22, 2)
    result["dose_young_ml"] = _cell_float(ws, 22, 3)

    wb.close()
    return {k: v for k, v in result.items() if v is not None}


def _map_procedure_type(raw: str) -> str:
    """Map Russian text to procedure_type enum."""
    raw = raw.lower()
    if "вакцин" in raw or "егіп" in raw:
        return "vaccination"
    if "аллерг" in raw or "туберкул" in raw:
        return "allergy_test"
    if "дегельм" in raw:
        return "deworming"
    if "лечен" in raw or "емдеу" in raw:
        return "treatment"
    return "vaccination"  # default


# ─── Template 2: Expendable materials (Без названия (3).xls) ─────────────────

def parse_materials_template(file_path: str | Path) -> list[dict]:
    """
    Parse expendable materials table.

    Expected layout:
    Row 1: Header (Наименование | Количество | Ед. изм.)
    Row 2+: Data rows

    Returns list of dicts: [{name, quantity, unit}, ...]
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl required")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Template file not found: {file_path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    materials = []

    for row in range(2, ws.max_row + 1):
        name = _cell_str(ws, row, 1)
        qty_raw = ws.cell(row=row, column=2).value
        unit = _cell_str(ws, row, 3)

        if not name or name.lower() in ("наименование", "атауы", ""):
            continue

        try:
            quantity = float(str(qty_raw)) if qty_raw is not None else 0
        except (ValueError, TypeError):
            quantity = 0

        materials.append({"name": name, "quantity": quantity, "unit": unit or "шт"})

    wb.close()
    return materials


# ─── Template 3: Animal registry (Без названия (5).xls) ──────────────────────

def parse_animal_registry_template(file_path: str | Path) -> list[dict]:
    """
    Parse individual animal records from registry table.

    Expected layout:
    Row 1: Header (№ | Ном./ID | Пол | Возраст | Окрас | ...)
    Row 2+: Data rows

    Returns list of dicts matching ProcedureActAnimal fields.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl required")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Template file not found: {file_path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    animals = []

    for row in range(2, ws.max_row + 1):
        id_no = _cell_str(ws, row, 2)
        if not id_no:
            continue

        sex_raw = _cell_str(ws, row, 3).lower()
        sex = "male" if sex_raw in ("м", "male", "еркек") else \
              "female" if sex_raw in ("ж", "female", "ұрғашы") else None

        age_raw = _cell_str(ws, row, 4).lower()
        age_group = "adult" if age_raw in ("взр.", "adult", "ересек") else \
                    "young" if age_raw in ("мол.", "young", "жас") else None

        record = {
            "identification_no": id_no,
            "sex": sex,
            "age_group": age_group,
            "color": _cell_str(ws, row, 5) or None,
        }

        # Allergy test columns (if present: col 6=skin_mm, 7=result_mm)
        skin_mm = _cell_float(ws, row, 6)
        result_mm = _cell_float(ws, row, 7)
        if skin_mm is not None:
            record["skin_measurement_mm"] = skin_mm
        if result_mm is not None:
            record["result_mm"] = result_mm
            if skin_mm is not None:
                record["difference_mm"] = round(result_mm - skin_mm, 1)
                record["allergy_result"] = _classify_allergy(result_mm - skin_mm)

        # Notes (last column)
        notes = _cell_str(ws, row, 8)
        if notes:
            record["notes"] = notes

        animals.append(record)

    wb.close()
    return animals


def _classify_allergy(diff_mm: float) -> str:
    """Classify allergy test result by skin fold difference (mm)."""
    if diff_mm < 3:
        return "negative"
    if diff_mm < 8:
        return "doubtful"
    return "positive"


# ─── Template 4: Signature block (Без названия (6).xls) ──────────────────────

def parse_signature_template(file_path: str | Path) -> dict:
    """
    Parse owner signature block.
    Returns specialist name and owner name from signature rows.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl required")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Template file not found: {file_path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    result = {
        "specialist_name": _cell_str(ws, 2, 1) or None,
        "owner_name": _cell_str(ws, 2, 3) or None,
    }

    wb.close()
    return {k: v for k, v in result.items() if v is not None}


# ─── Combined parser ──────────────────────────────────────────────────────────

def parse_full_act_from_templates(
    main_template: str | Path,
    materials_template: Optional[str | Path] = None,
    animal_registry_template: Optional[str | Path] = None,
    signature_template: Optional[str | Path] = None,
) -> dict:
    """
    Parse all 4 Excel templates and merge into a single procedure act dict.

    Returns a dict that can be passed directly to the procedure act API.
    """
    result = {}

    # Main act data
    try:
        result.update(parse_main_act_template(main_template))
    except Exception as e:
        logger.error(f"Failed to parse main act template: {e}")
        raise

    # Materials
    if materials_template:
        try:
            import json
            materials = parse_materials_template(materials_template)
            if materials:
                result["materials_json"] = json.dumps(materials, ensure_ascii=False)
        except Exception as e:
            logger.warning(f"Failed to parse materials template: {e}")

    # Animal registry
    if animal_registry_template:
        try:
            animals = parse_animal_registry_template(animal_registry_template)
            if animals:
                result["animals"] = animals
        except Exception as e:
            logger.warning(f"Failed to parse animal registry template: {e}")

    # Signature block
    if signature_template:
        try:
            sig_data = parse_signature_template(signature_template)
            # Only override if not already set
            result.setdefault("specialist_name", sig_data.get("specialist_name"))
            result.setdefault("owner_name", sig_data.get("owner_name"))
        except Exception as e:
            logger.warning(f"Failed to parse signature template: {e}")

    return result


# ─── XLS support (legacy .xls files via xlrd) ────────────────────────────────

def read_xls_sheet(file_path: str | Path) -> list[list]:
    """
    Read .xls file and return list of rows (list of cell values).
    Uses xlrd library.
    """
    if not XLRD_AVAILABLE:
        raise RuntimeError("xlrd is required for .xls parsing. Install: pip install xlrd==1.2.0")

    import xlrd
    path = str(file_path)
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)

    rows = []
    for row_idx in range(ws.nrows):
        row = []
        for col_idx in range(ws.ncols):
            cell = ws.cell(row_idx, col_idx)
            # Handle date cells (xlrd type 3)
            if cell.ctype == 3:
                try:
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row.append(dt)
                except Exception:
                    row.append(cell.value)
            else:
                row.append(cell.value)
        rows.append(row)

    return rows
